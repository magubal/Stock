"""Write all 6 stocks with new format: 해자(숫자), 투자가치(숫자), 비고(상세)"""
import sys
sys.stdout.reconfigure(encoding='utf-8')
from datetime import date
import openpyxl

XLSX_PATH = r"f:\PSJ\AntigravityWorkPlace\Stock\Test_02\data\국내상장종목 해자 투자가치.xlsx"
OUT_PATH  = r"f:\PSJ\AntigravityWorkPlace\Stock\Test_02\data\국내상장종목 해자 투자가치_결과.xlsx"

today = date.today().isoformat()

results = {
    "003610": {  # 방림
        "평가일자": today, "해자": 1, "투자가치": 0,
        "BM": "섬유제조(방적/직물)",
        "비고": (
            "해자 1/5: 증거15건 | 전환비용(5)+특허공정(4)+브랜드(3)+규모경제(3) | CAGR -7.7% 마진Δ-2.5%p 구조적악화 성장-1감점\n"
            "투자가치 0/5: 수익성0(영업적자) | 해자0.3 | 밸류0(산출불가) | TTM매출1,180억 영업이익-2억(-0.2%) 시총1,968억 | [oracle/high 20252Q]"
        ),
    },
    "005900": {  # 동양건설산업
        "평가일자": today, "해자": 2, "투자가치": 0,
        "BM": "토목건설(도로/지하철/항만)",
        "비고": (
            "해자 2/5: 증거5건 | 규모경제(2)+브랜드(3) | BM완성도17%\n"
            "투자가치 0/5: 수익성0(영업적자) | 해자0.6 | 밸류0(산출불가) | TTM매출1,307억 영업이익-400억(-30.6%) | Oracle TTM 20133Q 오래된데이터주의 | [oracle/high]"
        ),
    },
    "049130": {  # 하우리
        "평가일자": today, "해자": 2, "투자가치": 1,
        "BM": "보안SW(안티바이러스)",
        "비고": (
            "해자 2/5: 증거10건 | 전환비용(10) 조달청계약기반 | CAGR 69.6%↑ 마진Δ-19.7%p↓ 양질불일치\n"
            "투자가치 1/5: 수익성1.5(마진27.2%) | 해자0.6 | 밸류0(시총조회실패) | TTM매출92억 영업이익25억 | BM완성도17% 감점-0.5 | [oracle/high 20243Q]"
        ),
    },
    "008120": {  # 엠소닉
        "평가일자": today, "해자": 3, "투자가치": 1,
        "BM": "전자부품/스피커(제조)",
        "비고": (
            "해자 3/5: 증거13건 | 전환비용(8)+브랜드(3)+특허공정(2) | CAGR 14.7% 마진Δ+1.3%p 성장+1가점\n"
            "투자가치 1/5: 수익성0.5(마진4.4%) | 해자0.9 | 밸류0(시총조회실패) | TTM매출411억 영업이익18억 | [oracle/high 20243Q]"
        ),
    },
    "008340": {  # 대주코레스
        "평가일자": today, "해자": 3, "투자가치": 1,
        "BM": "알루미늄압출(자동차부품/건자재)",
        "비고": (
            "해자 3/5: 증거8건 | 전환비용(3)+브랜드(3)+규모경제(2) | CAGR 93.4%↑ 마진Δ-6.3%p↓ 양질불일치\n"
            "투자가치 1/5: 수익성0.5(마진3.5%) | 해자0.9 | 밸류0(시총조회실패) | TTM매출1,964억 영업이익68억 | [oracle/high 20243Q]"
        ),
    },
    "008870": {  # 금비
        "평가일자": today, "해자": 2, "투자가치": 2,
        "BM": "유리용기/화장품용기(제조)",
        "비고": (
            "해자 2/5: 증거27건 | 전환비용(9)+규모경제(6)+특허공정(5)+네트워크효과(2)+원가우위(2)+브랜드(2)+규제허가(1) | CAGR 2.7% 저성장\n"
            "투자가치 2/5: 수익성0.5(마진1.4%) | 해자0.6 | 밸류0.75(op_multiple 12.5x) | TTM매출2,524억 영업이익35억 시총437억 현재가49,500원 | [oracle/high 20253Q]"
        ),
    },
}

wb = openpyxl.load_workbook(XLSX_PATH)
ws = wb.active

headers = {}
for c in range(1, ws.max_column + 1):
    val = ws.cell(row=1, column=c).value
    if val:
        headers[val] = c

if "비고" not in headers:
    new_col = ws.max_column + 1
    ws.cell(row=1, column=new_col).value = "비고"
    headers["비고"] = new_col
    print(f"Added '비고' column at position {new_col}")

col_code = headers["종목코드"]
updated = 0
for row_idx in range(2, ws.max_row + 1):
    ticker = str(ws.cell(row=row_idx, column=col_code).value or "").strip()
    if ticker in results:
        r = results[ticker]
        ws.cell(row=row_idx, column=headers["평가일자"]).value = r["평가일자"]
        ws.cell(row=row_idx, column=headers["해자"]).value = r["해자"]
        ws.cell(row=row_idx, column=headers["투자가치"]).value = r["투자가치"]
        ws.cell(row=row_idx, column=headers["BM"]).value = r["BM"]
        ws.cell(row=row_idx, column=headers["비고"]).value = r["비고"]
        ws.cell(row=row_idx, column=headers["비고"]).alignment = openpyxl.styles.Alignment(wrap_text=True, vertical="top")
        updated += 1
        name = ws.cell(row=row_idx, column=2).value
        print(f"  Row {row_idx}: {ticker} {name} | 해자={r['해자']} 투자가치={r['투자가치']}")

wb.save(OUT_PATH)
print(f"\nDone: {updated} stocks -> {OUT_PATH}")
