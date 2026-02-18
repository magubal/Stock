"""
배치 해자/투자가치 분석 러너

엑셀에서 미평가 종목을 읽어 → 분석 → 결과를 바로 엑셀에 기록.
한 번 실행하면 끝까지 자동으로 돌아감 (중간 승인 불필요).

Usage:
    # 미평가 종목 중 앞에서 N개 분석
    python scripts/stock_moat/batch_moat_value.py --limit 10

    # 특정 행부터 재개 (6번째 행부터)
    python scripts/stock_moat/batch_moat_value.py --start-row 6 --limit 5

    # 전체 재분석 (이미 평가된 것도 덮어쓰기)
    python scripts/stock_moat/batch_moat_value.py --force --limit 20

    # 다른 엑셀 파일 지정
    python scripts/stock_moat/batch_moat_value.py --file data/my_stocks.xlsx --limit 5
"""

import sys
import os
import argparse
import time
from datetime import date
from pathlib import Path

if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8")

# Add paths
project_root = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(project_root / ".agent/skills/stock-moat/utils"))
sys.path.insert(0, str(project_root / "scripts/stock_moat"))
sys.path.insert(0, str(project_root / "scripts/idea_pipeline"))

import openpyxl
from openpyxl.styles import Alignment
from data_quality import format_krw
try:
    from monitoring_adapter import enforce_monitoring
    from scripts.consistency.fail_closed_runtime import register_fail_closed_guard, mark_monitoring_called
    HAS_MONITORING = True
except ImportError:
    HAS_MONITORING = False


DEFAULT_XLSX = str(project_root / "data" / "국내상장종목 해자 투자가치.xlsx")
REQUIRED_HEADERS = ["종목코드", "종목명", "평가일자", "투자가치", "BM", "해자"]


def load_workbook_safe(path: str):
    """엑셀 열기. 잠금 파일 있으면 경고."""
    xlsx_path = Path(path)
    lock_file = xlsx_path.parent / f"~${xlsx_path.name}"
    if lock_file.exists():
        print(f"  ⚠️ 엑셀이 열려있음 (잠금파일 감지). 결과를 _결과 파일에 저장합니다.")
        return None, True
    return openpyxl.load_workbook(path), False


def find_or_create_column(ws, header_name: str) -> int:
    """헤더가 없으면 새 컬럼 추가."""
    for c in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=c).value == header_name:
            return c
    new_col = ws.max_column + 1
    ws.cell(row=1, column=new_col).value = header_name
    return new_col


def build_bigo(result: dict) -> str:
    """비고 컬럼 텍스트 (해자 상세 + 투자가치 상세, 각 1줄)."""
    # 해자 라인
    moat = result.get("해자강도", 0)
    ev_count = result.get("evidence_count", 0)
    ev_quality = result.get("evidence_quality", 0)
    growth_adj = result.get("growth_adjustment", 0)
    growth_reason = result.get("growth_reason", "")
    cagr = result.get("revenue_cagr", "")
    margin_d = result.get("op_margin_delta", "")

    moat_parts = [f"해자 {moat}/5", f"증거{ev_count}건(q{ev_quality:.0f})"]
    if growth_adj != 0:
        moat_parts.append(f"성장{growth_adj:+d}")
    if growth_reason:
        moat_parts.append(growth_reason)

    # 투자가치 라인
    iv = result.get("investment_value", 0)
    iv_reason = result.get("investment_value_reason", "")

    ttm_rev = result.get("ttm_revenue", 0)
    ttm_op = result.get("ttm_op_income", 0)
    ttm_margin = result.get("ttm_op_margin", "")
    op_mult = result.get("op_multiple")
    market_cap = result.get("market_cap", 0)
    price = result.get("current_price", 0)
    source = result.get("data_source", "")
    confidence = result.get("data_confidence", "")
    ttm_q = result.get("ttm_quarter", "")

    val_parts = [f"투자가치 {iv}/5"]
    if ttm_rev:
        val_parts.append(f"TTM매출{format_krw(ttm_rev)}")
    if ttm_op:
        val_parts.append(f"영업이익{format_krw(ttm_op)}({ttm_margin})")
    if op_mult is not None:
        val_parts.append(f"op_multiple {op_mult:.1f}x")
    elif ttm_op <= 0:
        val_parts.append("op_multiple N/A(적자)")
    if market_cap:
        val_parts.append(f"시총{format_krw(market_cap)}")
    if price:
        val_parts.append(f"현재가{price:,.0f}원")
    val_parts.append(f"[{source}/{confidence} {ttm_q}]")

    return " | ".join(moat_parts) + "\n" + " | ".join(val_parts)


def build_bm_short(result: dict) -> str:
    """BM 컬럼 (간략 1줄)."""
    sector_top = result.get("core_sector_top", "")
    sector_sub = result.get("core_sector_sub", "")
    if sector_top and sector_sub and sector_top != sector_sub:
        return f"{sector_top}/{sector_sub}"
    return sector_top or sector_sub or ""


def main():
    parser = argparse.ArgumentParser(description="배치 해자/투자가치 분석")
    parser.add_argument("--file", type=str, default=DEFAULT_XLSX, help="엑셀 파일 경로")
    parser.add_argument("--limit", type=int, default=10, help="분석할 종목 수 (기본 10)")
    parser.add_argument("--start-row", type=int, default=0, help="시작 행 번호 (2=첫 데이터행)")
    parser.add_argument("--force", action="store_true", help="이미 평가된 종목도 재분석")
    parser.add_argument("--year", type=str, default="auto", help="보고서 기준년도")
    parser.add_argument("--save-every", type=int, default=10, help="N종목마다 중간저장 (기본 10)")
    args = parser.parse_args()
    if HAS_MONITORING:
        register_fail_closed_guard("scripts.stock_moat.batch_moat_value")

    today = date.today().isoformat()

    # ── 1. 엑셀 읽기 ──
    print(f"\n{'='*70}")
    print(f"  배치 해자/투자가치 분석")
    print(f"  파일: {args.file}")
    print(f"  설정: limit={args.limit}, start-row={args.start_row}, force={args.force}")
    print(f"{'='*70}")

    wb, is_locked = load_workbook_safe(args.file)
    if wb is None:
        # 잠금 시 원본 읽고 다른 파일에 저장
        wb = openpyxl.load_workbook(args.file)
    ws = wb.active

    # 헤더 매핑
    headers = {}
    for c in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=c).value
        if val:
            headers[val] = c

    col_code = headers.get("종목코드")
    col_name = headers.get("종목명")
    col_date = headers.get("평가일자")
    col_value = headers.get("투자가치")
    col_bm = headers.get("BM")
    col_moat = headers.get("해자")
    col_bigo = find_or_create_column(ws, "비고")

    if not col_code or not col_name:
        if HAS_MONITORING:
            mark_monitoring_called()
        print("❌ 종목코드/종목명 컬럼을 찾을 수 없습니다.")
        return

    # ── 2. 분석 대상 수집 ──
    targets = []
    for row_idx in range(2, ws.max_row + 1):
        ticker = str(ws.cell(row=row_idx, column=col_code).value or "").strip()
        name = str(ws.cell(row=row_idx, column=col_name).value or "").strip()
        existing_moat = ws.cell(row=row_idx, column=col_moat).value

        if not ticker or not name:
            continue

        if args.start_row and row_idx < args.start_row:
            continue

        if not args.force and existing_moat is not None:
            continue  # 이미 평가됨 → 스킵

        targets.append((row_idx, ticker, name))

        if len(targets) >= args.limit:
            break

    print(f"\n  분석 대상: {len(targets)}개 종목")
    if not targets:
        if HAS_MONITORING:
            mark_monitoring_called()
        print("  ✅ 분석할 종목이 없습니다.")
        return

    # ── 3. 파이프라인 초기화 (1회만) ──
    from analyze_with_evidence import EvidenceBasedMoatPipeline
    pipeline = EvidenceBasedMoatPipeline()

    # ── 4. 순차 분석 + 엑셀 기록 ──
    success_count = 0
    fail_count = 0

    for i, (row_idx, ticker, name) in enumerate(targets):
        print(f"\n[{i+1}/{len(targets)}] {name} ({ticker}) — Row {row_idx}")

        try:
            result = pipeline.analyze_stock(ticker, name, args.year)

            # Optional monitoring guard
            if HAS_MONITORING:
                monitor_result = enforce_monitoring(
                    {
                        "req_id": "REQ-007",
                        "enforce_requirement_contract": True,
                        "consistency_monitoring_enabled": True,
                        "source_path": "scripts.stock_moat.batch_moat_value",
                        "entity_type": "stock_analysis",
                        "entity_id": f"{ticker}:{today}",
                        "requirement_refs": ["REQUESTS.md#REQ-007", "REQUESTS.md#REQ-008", "REQUESTS.md#REQ-009"],
                        "plan_refs": ["docs/plans/2026-02-15-global-monitoring-guard-implementation.md"],
                        "design_refs": [
                            "docs/plans/2026-02-15-global-monitoring-guard-design.md",
                            "docs/plans/2026-02-15-global-monitoring-guard-implementation.md",
                        ],
                        "test_tags": ["stock_batch", "monitoring_guard"],
                        "category": "PORTFOLIO",
                        "packet_type": "종목",
                        "stock_context": True,
                        "run_stock_pipeline": True,
                        "pipeline_executed": result.get("status") != "failed",
                        "pipeline_error": result.get("error", ""),
                        "create_idea": False,
                        "idea_gate_should_create": False,
                        "force_create_idea": False,
                        "traceability_ok": bool(result.get("bm_summary") or result.get("core_desc")),
                        "ticker": ticker,
                        "stock_name": name,
                    },
                    hard_block=True,
                )
                if monitor_result.blocked:
                    fail_count += 1
                    ws.cell(row=row_idx, column=col_date).value = today
                    ws.cell(row=row_idx, column=col_moat).value = result.get("해자강도", 0)
                    ws.cell(row=row_idx, column=col_value).value = result.get("investment_value", 0)
                    ws.cell(row=row_idx, column=col_bm).value = "MONITOR_BLOCK"
                    ws.cell(row=row_idx, column=col_bigo).value = (
                        f"MONITOR_BLOCK[{monitor_result.rule_code}] "
                        f"incident={monitor_result.incident_id} "
                        f"reasons={','.join(monitor_result.reasons or [])}"
                    )
                    print(
                        f"  [MONITOR BLOCK] {monitor_result.rule_code} "
                        f"(incident={monitor_result.incident_id})"
                    )
                    continue

            if result.get("status") == "failed":
                print(f"  ❌ 실패: {result.get('error')}")
                fail_count += 1
                # 실패해도 최소한 기록
                ws.cell(row=row_idx, column=col_date).value = today
                ws.cell(row=row_idx, column=col_moat).value = result.get("해자강도", 0)
                ws.cell(row=row_idx, column=col_value).value = 0
                ws.cell(row=row_idx, column=col_bm).value = "분석실패"
                ws.cell(row=row_idx, column=col_bigo).value = f"ERROR: {result.get('error', 'unknown')}"
                continue

            # 성공 → 엑셀 기록
            ws.cell(row=row_idx, column=col_date).value = today
            ws.cell(row=row_idx, column=col_moat).value = result.get("해자강도", 0)
            ws.cell(row=row_idx, column=col_value).value = result.get("investment_value", 0)
            ws.cell(row=row_idx, column=col_bm).value = build_bm_short(result)
            ws.cell(row=row_idx, column=col_bigo).value = build_bigo(result)
            ws.cell(row=row_idx, column=col_bigo).alignment = Alignment(wrap_text=True, vertical="top")

            success_count += 1
            print(f"  ✅ 해자={result.get('해자강도')}/5, 투자가치={result.get('investment_value')}/5")

        except Exception as e:
            print(f"  ❌ 예외: {e}")
            fail_count += 1
            ws.cell(row=row_idx, column=col_date).value = today
            ws.cell(row=row_idx, column=col_bigo).value = f"EXCEPTION: {e}"

        # Rate limiting
        if i < len(targets) - 1:
            time.sleep(1)

        # ── 중간저장 (N종목마다) ──
        if (i + 1) % args.save_every == 0:
            _save_path = args.file.replace(".xlsx", "_결과.xlsx") if is_locked else args.file
            try:
                wb.save(_save_path)
                print(f"\n  💾 중간저장 ({i+1}/{len(targets)}) → {Path(_save_path).name}")
            except PermissionError:
                _save_path = args.file.replace(".xlsx", "_결과.xlsx")
                wb.save(_save_path)
                print(f"\n  💾 중간저장 ({i+1}/{len(targets)}) → {Path(_save_path).name} (잠금우회)")

    # ── 5. 최종 저장 ──
    if is_locked:
        out_path = args.file.replace(".xlsx", "_결과.xlsx")
    else:
        out_path = args.file

    try:
        wb.save(out_path)
        print(f"\n✅ 최종 저장 완료: {out_path}")
    except PermissionError:
        out_path = args.file.replace(".xlsx", "_결과.xlsx")
        wb.save(out_path)
        print(f"\n⚠️ 원본 잠금 → 별도 저장: {out_path}")

    # ── 6. 요약 ──
    print(f"\n{'='*70}")
    print(f"  배치 완료: 성공 {success_count}, 실패 {fail_count}")
    print(f"  결과 파일: {out_path}")
    print(f"{'='*70}")


if __name__ == "__main__":
    main()
