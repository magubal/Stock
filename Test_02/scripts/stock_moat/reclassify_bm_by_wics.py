"""
Reclassify BM column using Naver Finance WICS labels.

Source page:
  https://navercomp.wisereport.co.kr/v2/company/c1010001.aspx?cmp_cd={ticker}
"""

import argparse
import json
import re
import shutil
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from pathlib import Path
from typing import Dict, Optional, Tuple

import openpyxl
import requests


PROJECT_ROOT = Path(__file__).resolve().parents[2]
DATA_DIR = PROJECT_ROOT / "data"
CACHE_PATH = DATA_DIR / "wics_cache.json"
WICS_URL = "https://navercomp.wisereport.co.kr/v2/company/c1010001.aspx?cmp_cd={ticker}"
USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0 Safari/537.36"
)

_THREAD_LOCAL = threading.local()


def extract_wics_from_html(html: str) -> Optional[str]:
    match = re.search(r"WICS\s*:\s*([^<\r\n]+)", html, flags=re.IGNORECASE)
    if not match:
        return None
    value = match.group(1).strip()
    return value or None


def _get_session() -> requests.Session:
    session = getattr(_THREAD_LOCAL, "session", None)
    if session is None:
        session = requests.Session()
        session.headers.update({"User-Agent": USER_AGENT})
        _THREAD_LOCAL.session = session
    return session


def fetch_wics(ticker: str, timeout: float = 12.0, retries: int = 3) -> Tuple[Optional[str], Optional[str]]:
    url = WICS_URL.format(ticker=ticker)
    last_error = None
    for _ in range(retries):
        try:
            resp = _get_session().get(url, timeout=timeout)
            if resp.status_code != 200:
                last_error = f"http_{resp.status_code}"
                continue
            wics = extract_wics_from_html(resp.text)
            if wics:
                return wics, None
            last_error = "wics_not_found"
        except Exception as exc:  # noqa: BLE001
            last_error = f"request_error:{exc.__class__.__name__}"
    return None, last_error


def load_cache() -> Dict[str, Dict[str, str]]:
    if not CACHE_PATH.exists():
        return {}
    try:
        payload = json.loads(CACHE_PATH.read_text(encoding="utf-8"))
        return payload.get("items", {})
    except Exception:  # noqa: BLE001
        return {}


def save_cache(items: Dict[str, Dict[str, str]]) -> None:
    payload = {
        "updated_at": datetime.now().isoformat(),
        "items": items,
    }
    CACHE_PATH.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def resolve_workbook_path(arg_path: Optional[str]) -> Path:
    if arg_path:
        return Path(arg_path).resolve()

    candidates = [
        p
        for p in DATA_DIR.glob("*.xlsx")
        if "_결과" not in p.name and "backup" not in p.name
    ]
    if not candidates:
        raise FileNotFoundError("No workbook found under data/*.xlsx")
    candidates.sort(key=lambda p: p.stat().st_size, reverse=True)
    return candidates[0]


def main() -> None:
    parser = argparse.ArgumentParser(description="Reclassify BM by WICS from Naver Finance")
    parser.add_argument("--file", type=str, default=None, help="Workbook path")
    parser.add_argument("--workers", type=int, default=12, help="Parallel worker count")
    parser.add_argument("--timeout", type=float, default=12.0, help="HTTP timeout")
    parser.add_argument("--force-refresh", action="store_true", help="Ignore cache and refetch all")
    parser.add_argument("--dry-run", action="store_true", help="Compute only, do not save workbook")
    args = parser.parse_args()

    workbook_path = resolve_workbook_path(args.file)
    if not workbook_path.exists():
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")

    print(f"[WICS] Workbook: {workbook_path}")
    wb = openpyxl.load_workbook(workbook_path)
    ws = wb.active

    # Fixed schema in this project:
    # 1 code, 2 name, 5 BM, 6 moat
    col_code = 1
    col_name = 2
    col_bm = 5
    col_moat = 6

    targets = []
    for row_idx in range(2, ws.max_row + 1):
        moat = ws.cell(row_idx, col_moat).value
        if moat is None:
            continue
        ticker = str(ws.cell(row_idx, col_code).value or "").strip().zfill(6)
        if not ticker:
            continue
        name = str(ws.cell(row_idx, col_name).value or "").strip()
        old_bm = str(ws.cell(row_idx, col_bm).value or "").strip()
        targets.append((row_idx, ticker, name, old_bm))

    print(f"[WICS] Targets (evaluated rows): {len(targets)}")
    cache = load_cache()
    now = datetime.now().isoformat()

    resolved: Dict[str, Dict[str, Optional[str]]] = {}
    missing = []
    for _, ticker, _, _ in targets:
        cached = cache.get(ticker)
        if cached and cached.get("wics") and not args.force_refresh:
            resolved[ticker] = {"wics": cached["wics"], "error": None}
        else:
            missing.append(ticker)

    print(f"[WICS] Cache hit: {len(resolved)} / Fetch needed: {len(missing)}")

    if missing:
        with ThreadPoolExecutor(max_workers=max(1, args.workers)) as executor:
            future_map = {
                executor.submit(fetch_wics, ticker, args.timeout, 3): ticker
                for ticker in missing
            }
            done = 0
            total = len(missing)
            for future in as_completed(future_map):
                ticker = future_map[future]
                wics, error = future.result()
                resolved[ticker] = {"wics": wics, "error": error}
                done += 1
                if done % 200 == 0 or done == total:
                    print(f"[WICS] Fetch progress: {done}/{total}")

    updated = 0
    unchanged = 0
    failed = []
    for row_idx, ticker, name, old_bm in targets:
        item = resolved.get(ticker, {})
        wics = item.get("wics")
        err = item.get("error")
        if not wics:
            failed.append({"ticker": ticker, "name": name, "error": err})
            unchanged += 1
            continue
        if old_bm != wics:
            ws.cell(row_idx, col_bm).value = wics
            updated += 1
        else:
            unchanged += 1
        cache[ticker] = {"wics": wics, "fetched_at": now}

    print(f"[WICS] Updated BM rows: {updated}")
    print(f"[WICS] Unchanged rows: {unchanged}")
    print(f"[WICS] Fetch failures: {len(failed)}")

    if failed:
        fail_path = DATA_DIR / f"wics_fetch_failures_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        fail_path.write_text(json.dumps(failed, ensure_ascii=False, indent=2), encoding="utf-8")
        print(f"[WICS] Failure list: {fail_path}")

    if not args.dry_run:
        backup_path = workbook_path.with_name(
            f"{workbook_path.stem}_wics_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}{workbook_path.suffix}"
        )
        shutil.copy2(workbook_path, backup_path)
        wb.save(workbook_path)
        print(f"[WICS] Backup: {backup_path}")
        print(f"[WICS] Saved workbook: {workbook_path}")
    else:
        print("[WICS] Dry-run mode; workbook not saved")

    wb.close()
    save_cache(cache)
    print(f"[WICS] Cache saved: {CACHE_PATH}")


if __name__ == "__main__":
    main()
