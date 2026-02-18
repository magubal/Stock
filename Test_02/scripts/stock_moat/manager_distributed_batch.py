"""
Manager-distributed batch runner for moat/investment-value analysis.

Design:
- Manager splits target rows into contiguous shards.
- Each shard is processed by a worker process running `batch_moat_value.py`
  against its own copied workbook.
- Manager merges evaluated columns back into a single output workbook.

This keeps analysis logic identical to existing worker script while enabling
parallel execution with isolated files.
"""

import argparse
import shutil
import subprocess
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Tuple

import openpyxl


HEADER_CODE = "종목코드"
HEADER_NAME = "종목명"
HEADER_DATE = "평가일자"
HEADER_VALUE = "투자가치"
HEADER_BM = "BM"
HEADER_MOAT = "해자"
HEADER_BIGO = "비고"

COPY_HEADERS = [
    HEADER_DATE,
    HEADER_VALUE,
    HEADER_BM,
    HEADER_MOAT,
    HEADER_BIGO,
]


@dataclass
class TargetRow:
    row_idx: int
    ticker: str
    name: str


@dataclass
class WorkerJob:
    worker_id: int
    workbook_path: Path
    start_row: int
    limit: int
    targets: List[TargetRow]


def find_headers(ws) -> Dict[str, int]:
    headers: Dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if v is not None:
            headers[str(v).strip()] = c
    return headers


def ensure_header(ws, header_name: str, headers: Dict[str, int]) -> int:
    if header_name in headers:
        return headers[header_name]
    col = ws.max_column + 1
    ws.cell(row=1, column=col).value = header_name
    headers[header_name] = col
    return col


def collect_targets(
    ws,
    headers: Dict[str, int],
    start_row: int,
    limit: int,
    force: bool,
) -> List[TargetRow]:
    col_code = headers.get(HEADER_CODE)
    col_name = headers.get(HEADER_NAME)
    col_moat = headers.get(HEADER_MOAT)

    if not col_code or not col_name:
        raise ValueError("종목코드/종목명 헤더를 찾을 수 없습니다.")

    targets: List[TargetRow] = []
    row_begin = max(2, start_row if start_row > 0 else 2)

    for r in range(row_begin, ws.max_row + 1):
        ticker = str(ws.cell(row=r, column=col_code).value or "").strip()
        name = str(ws.cell(row=r, column=col_name).value or "").strip()
        existing_moat = ws.cell(row=r, column=col_moat).value if col_moat else None

        if not ticker or not name:
            continue
        if (not force) and (existing_moat is not None):
            continue

        targets.append(TargetRow(row_idx=r, ticker=ticker, name=name))
        if limit > 0 and len(targets) >= limit:
            break

    return targets


def split_contiguous(targets: List[TargetRow], workers: int) -> List[List[TargetRow]]:
    if workers <= 1 or len(targets) <= 1:
        return [targets]

    workers = max(1, min(workers, len(targets)))
    base = len(targets) // workers
    rem = len(targets) % workers

    chunks: List[List[TargetRow]] = []
    i = 0
    for w in range(workers):
        n = base + (1 if w < rem else 0)
        if n <= 0:
            continue
        chunks.append(targets[i:i + n])
        i += n
    return chunks


def build_worker_jobs(
    source_file: Path,
    chunks: List[List[TargetRow]],
    work_dir: Path,
) -> List[WorkerJob]:
    jobs: List[WorkerJob] = []
    stem = source_file.stem
    suffix = source_file.suffix

    for idx, chunk in enumerate(chunks, start=1):
        worker_file = work_dir / f"{stem}.worker{idx}{suffix}"
        shutil.copy2(source_file, worker_file)

        jobs.append(
            WorkerJob(
                worker_id=idx,
                workbook_path=worker_file,
                start_row=chunk[0].row_idx,
                limit=len(chunk),
                targets=chunk,
            )
        )

    return jobs


def launch_workers(
    jobs: List[WorkerJob],
    year: str,
    force: bool,
    python_exe: str,
    worker_script: Path,
) -> List[Tuple[WorkerJob, int, str, str]]:
    procs: List[Tuple[WorkerJob, subprocess.Popen]] = []

    for job in jobs:
        cmd = [
            python_exe,
            str(worker_script),
            "--file",
            str(job.workbook_path),
            "--start-row",
            str(job.start_row),
            "--limit",
            str(job.limit),
            "--year",
            year,
        ]
        if force:
            cmd.append("--force")

        p = subprocess.Popen(
            cmd,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
            encoding="utf-8",
            errors="replace",
        )
        procs.append((job, p))

    results: List[Tuple[WorkerJob, int, str, str]] = []
    for job, p in procs:
        out, err = p.communicate()
        results.append((job, p.returncode, out, err))
    return results


def merge_results(
    source_file: Path,
    output_file: Path,
    jobs: List[WorkerJob],
) -> None:
    wb = openpyxl.load_workbook(source_file)
    ws = wb.active
    headers = find_headers(ws)

    for h in COPY_HEADERS:
        ensure_header(ws, h, headers)

    for job in jobs:
        wb_worker = openpyxl.load_workbook(job.workbook_path, data_only=False)
        ws_worker = wb_worker.active
        worker_headers = find_headers(ws_worker)

        # Ensure headers exist in worker too (for newly created 비고).
        if HEADER_BIGO not in worker_headers:
            worker_headers[HEADER_BIGO] = ws_worker.max_column + 1

        for target in job.targets:
            r = target.row_idx
            for h in COPY_HEADERS:
                src_col = worker_headers.get(h)
                dst_col = headers.get(h)
                if src_col is None or dst_col is None:
                    continue
                ws.cell(row=r, column=dst_col).value = ws_worker.cell(row=r, column=src_col).value

    wb.save(output_file)


def main() -> None:
    parser = argparse.ArgumentParser(description="Manager distributed moat batch runner")
    parser.add_argument(
        "--file",
        type=str,
        default=str(Path(__file__).resolve().parents[2] / "data" / "국내상장종목 해자 투자가치.xlsx"),
        help="입력 대상 엑셀 파일",
    )
    parser.add_argument("--workers", type=int, default=3, help="워커 수")
    parser.add_argument("--limit", type=int, default=0, help="분석 건수 제한(0=가능한 전부)")
    parser.add_argument("--start-row", type=int, default=0, help="시작 행 번호(2부터)")
    parser.add_argument("--force", action="store_true", help="기존 평가된 종목도 재분석")
    parser.add_argument("--year", type=str, default="auto", help="보고서 연도")
    parser.add_argument(
        "--output",
        type=str,
        default="",
        help="출력 파일 경로(미지정 시 *_manager_result.xlsx)",
    )
    parser.add_argument(
        "--keep-worker-files",
        action="store_true",
        help="워커 임시 파일 보관",
    )
    args = parser.parse_args()

    source_file = Path(args.file).resolve()
    if not source_file.exists():
        raise FileNotFoundError(f"입력 파일을 찾을 수 없습니다: {source_file}")

    if args.output:
        output_file = Path(args.output).resolve()
    else:
        output_file = source_file.with_name(f"{source_file.stem}_manager_result{source_file.suffix}")

    wb = openpyxl.load_workbook(source_file)
    ws = wb.active
    headers = find_headers(ws)
    targets = collect_targets(
        ws=ws,
        headers=headers,
        start_row=args.start_row,
        limit=args.limit,
        force=args.force,
    )
    wb.close()

    print(f"source={source_file}")
    print(f"targets={len(targets)} workers={args.workers} force={args.force}")
    if not targets:
        print("분석 대상이 없습니다.")
        return

    chunks = split_contiguous(targets, args.workers)
    print(f"chunks={len(chunks)}")

    worker_dir = source_file.parent / "_manager_workers"
    worker_dir.mkdir(parents=True, exist_ok=True)
    jobs = build_worker_jobs(source_file=source_file, chunks=chunks, work_dir=worker_dir)

    worker_script = Path(__file__).resolve().with_name("batch_moat_value.py")
    run_results = launch_workers(
        jobs=jobs,
        year=args.year,
        force=args.force,
        python_exe=sys.executable,
        worker_script=worker_script,
    )

    failed = []
    for job, code, out, err in run_results:
        print(f"[worker-{job.worker_id}] return={code} start_row={job.start_row} limit={job.limit}")
        if code != 0:
            failed.append((job.worker_id, code))
            if out.strip():
                print(f"[worker-{job.worker_id}] stdout tail:\n{out[-1200:]}")
            if err.strip():
                print(f"[worker-{job.worker_id}] stderr tail:\n{err[-1200:]}")

    if failed:
        raise RuntimeError(f"워커 실패 발생: {failed}")

    merge_results(source_file=source_file, output_file=output_file, jobs=jobs)
    print(f"merged_output={output_file}")

    if not args.keep_worker_files:
        for job in jobs:
            try:
                job.workbook_path.unlink(missing_ok=True)
            except Exception:
                pass


if __name__ == "__main__":
    main()

