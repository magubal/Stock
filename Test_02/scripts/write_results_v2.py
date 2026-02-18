"""Write analysis results with new format:
- 해자: numeric 0-5
- 투자가치: numeric 0-5
- 비고 (new column): detail text (해자 line + 투자가치 line)

투자가치 점수 기준 (0-5):
  0 = 구조적 적자 + 해자 없음
  1 = 영업적자 or 극히 낮은 수익성 + 해자 약함
  2 = 수익성 낮거나 불안정 + 해자 보통 이하
  3 = 보통 수익성 + 해자 보통
  4 = 양호한 수익성 + 해자 견고 + 합리적 밸류
  5 = 높은 수익성 + 해자 강함 + 매력적 밸류
"""
import sys
sys.stdout.reconfigure(encoding='utf-8')
from datetime import date
import openpyxl

XLSX_PATH = r"f:\PSJ\AntigravityWorkPlace\Stock\Test_02\data\국내상장종목 해자 투자가치.xlsx"
OUT_PATH  = r"f:\PSJ\AntigravityWorkPlace\Stock\Test_02\data\국내상장종목 해자 투자가치_결과.xlsx"

today = date.today().isoformat()

results = {
    "003610": {  # 방림
        "평가일자": today,
        "해자": 1,
        "투자가치": 0,
        "BM": "섬유제조(방적/직물)",
        "비고": (
            "해자: 1/5 | 증거15건 | 전환비용(5)+특허공정(4)+브랜드(3)+규모경제(3) | CAGR -7.7% 구조적악화 성장-1감점\n"
            "투자가치: 0/5 | TTM매출 1,180억 영업이익 -2억(-0.2%) 영업적자 | 시총 1,968억 | op_multiple N/A | [oracle/high 20252Q]"
        ),
    },
    "005900": {  # 동양건설산업
        "평가일자": today,
        "해자": 2,
        "투자가치": 0,
        "BM": "토목건설(도로/지하철/항만)",
        "비고": (
            "해자: 2/5 | 증거5건 | 규모경제(2)+브랜드(3) | BM완성도17% 정보부족\n"
            "투자가치: 0/5 | TTM매출 1,307억 영업이익 -400억(-30.6%) 심각적자 | Oracle TTM 20133Q 데이터오래됨 | [oracle/high]"
        ),
    },
    "049130": {  # 하우리
        "평가일자": today,
        "해자": 2,
        "투자가치": 2,
        "BM": "보안SW(안티바이러스)",
        "비고": (
            "해자: 2/5 | 증거10건 | 전환비용(10) 조달청계약기반 | CAGR 69.6%↑ 마진Δ-19.7%p↓ 질적성장 불확실\n"
            "투자가치: 2/5 | TTM매출 92억 영업이익 25억(27.2%) 소규모고마진 | 시총조회실패 | GICS:IT-GameSW(분류오류) | [oracle/high 20243Q]"
        ),
    },
    "008120": {  # 엠소닉
        "평가일자": today,
        "해자": 3,
        "투자가치": 2,
        "BM": "전자부품/스피커(제조)",
        "비고": (
            "해자: 3/5 | 증거13건 | 전환비용(8)+브랜드(3)+특허공정(2) | CAGR 14.7% 마진Δ+1.3%p 성장+1가점\n"
            "투자가치: 2/5 | TTM매출 411억 영업이익 18억(4.4%) 저마진 | 시총조회실패 | [oracle/high 20243Q]"
        ),
    },
}

wb = openpyxl.load_workbook(XLSX_PATH)
ws = wb.active

# Read existing headers
headers = {}
for c in range(1, ws.max_column + 1):
    val = ws.cell(row=1, column=c).value
    if val:
        headers[val] = c

# Add 비고 column if not exists
if "비고" not in headers:
    new_col = ws.max_column + 1
    ws.cell(row=1, column=new_col).value = "비고"
    headers["비고"] = new_col
    print(f"Added '비고' column at position {new_col}")

col_code = headers["종목코드"]

updated = 0
for row_idx in range(2, ws.max_row + 1):
    ticker = str(ws.cell(row=row_idx, column=col_code).value or "").strip()
    if ticker in results:
        r = results[ticker]
        ws.cell(row=row_idx, column=headers["평가일자"]).value = r["평가일자"]
        ws.cell(row=row_idx, column=headers["해자"]).value = r["해자"]
        ws.cell(row=row_idx, column=headers["투자가치"]).value = r["투자가치"]
        ws.cell(row=row_idx, column=headers["BM"]).value = r["BM"]
        ws.cell(row=row_idx, column=headers["비고"]).value = r["비고"]
        # Set wrap text for 비고 column
        ws.cell(row=row_idx, column=headers["비고"]).alignment = openpyxl.styles.Alignment(wrap_text=True, vertical="top")
        updated += 1
        name = ws.cell(row=row_idx, column=2).value
        print(f"  Row {row_idx}: {ticker} {name} | 해자={r['해자']} 투자가치={r['투자가치']}")

wb.save(OUT_PATH)
print(f"\nDone: {updated} stocks -> {OUT_PATH}")
