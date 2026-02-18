import sys
import json
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl

wb = openpyxl.load_workbook(r"f:\PSJ\AntigravityWorkPlace\Stock\Test_02\data\국내상장종목 해자 투자가치.xlsx")
ws = wb.active
print(f"Sheet: {ws.title}, Rows: {ws.max_row}, Cols: {ws.max_column}")

# Print headers
headers = [cell.value for cell in ws[1]]
print(f"Headers: {json.dumps(headers, ensure_ascii=False)}")

# Print first 5 data rows
for row_idx in range(2, min(7, ws.max_row + 1)):
    row_data = {headers[i]: ws.cell(row=row_idx, column=i+1).value for i in range(len(headers))}
    print(f"Row {row_idx}: {json.dumps(row_data, ensure_ascii=False, default=str)}")
