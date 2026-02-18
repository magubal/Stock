import sys, json
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
wb = openpyxl.load_workbook(r"f:\PSJ\AntigravityWorkPlace\Stock\Test_02\data\국내상장종목 해자 투자가치.xlsx")
ws = wb.active
for r in range(6, 8):
    code = ws.cell(row=r, column=1).value
    name = ws.cell(row=r, column=2).value
    print(f"Row {r}: {code} {name}")
