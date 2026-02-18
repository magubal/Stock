"""Write analysis results - save to new file to avoid lock."""
import sys
sys.stdout.reconfigure(encoding='utf-8')
from datetime import date
import openpyxl

XLSX_PATH = r"f:\PSJ\AntigravityWorkPlace\Stock\Test_02\data\국내상장종목 해자 투자가치.xlsx"
OUT_PATH  = r"f:\PSJ\AntigravityWorkPlace\Stock\Test_02\data\국내상장종목 해자 투자가치_결과.xlsx"

results = {
    "003610": {
        "평가일자": date.today().isoformat(),
        "투자가치": "op_multiple: N/A (영업적자) | TTM매출 1,180억 | TTM영업이익 -2억(-0.2%) | 시총 1,968억 | 성장CAGR -7.7% | [oracle/high]",
        "BM": "섬유제조(방적/직물) | 완성도 83% | 베트남 해외생산 | 군수용 NIR기술 | 전환비용(5)+특허공정(4)+브랜드(3)+규모경제(3) | GICS: Consumer Staples",
        "해자": "1/5 (구조적 악화: CAGR -7.7%, 마진Δ -2.5%p, 성장 -1 감점)",
    },
    "005900": {
        "평가일자": date.today().isoformat(),
        "투자가치": "op_multiple: N/A (영업적자) | TTM매출 1,307억 | TTM영업이익 -400억(-30.6%) | 시총 조회실패 | 성장CAGR 0% | [oracle/high]",
        "BM": "토목건설(도로/지하철/항만/고속철도) | 완성도 17% | 수주경쟁 치열 | 규모경제(2)+브랜드(3) | GICS: Industrials - Civil Engineering",
        "해자": "2/5 (증거 5건, 수익성 극히 악화, BM정보 부족)",
    },
}

wb = openpyxl.load_workbook(XLSX_PATH)
ws = wb.active

headers = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
col_code = headers["종목코드"]

updated = 0
for row_idx in range(2, ws.max_row + 1):
    ticker = str(ws.cell(row=row_idx, column=col_code).value or "").strip()
    if ticker in results:
        r = results[ticker]
        ws.cell(row=row_idx, column=headers["평가일자"]).value = r["평가일자"]
        ws.cell(row=row_idx, column=headers["투자가치"]).value = r["투자가치"]
        ws.cell(row=row_idx, column=headers["BM"]).value = r["BM"]
        ws.cell(row=row_idx, column=headers["해자"]).value = r["해자"]
        updated += 1
        print(f"  Updated row {row_idx}: {ticker} ({ws.cell(row=row_idx, column=2).value})")

wb.save(OUT_PATH)
print(f"\nDone: {updated} stocks updated -> {OUT_PATH}")
