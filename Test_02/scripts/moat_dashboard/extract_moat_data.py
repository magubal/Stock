"""
Extract moat analysis data from Excel to JSON for dashboard.
Reads: data/국내상장종목 해자 투자가치.xlsx
Writes: data/moat_analysis_summary.json
Also syncs Excel source data to DB:
  - moat_stock_snapshot (latest state)
  - moat_stock_history (change history)
  - moat_stock_sync_runs (sync run summary)
"""

from datetime import datetime
import hashlib
import json
import re
import sys
from pathlib import Path
from typing import Dict, List, Optional, Tuple

if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8")

import openpyxl
import requests
from sqlalchemy import inspect as sa_inspect, text as sa_text
from sqlalchemy.orm import Session

project_root = Path(__file__).resolve().parents[2]
XLSX_PATH = str(project_root / "data" / "국내상장종목 해자 투자가치.xlsx")
OUTPUT_PATH = str(project_root / "data" / "moat_analysis_summary.json")
DEFAULT_DB_PATH = project_root / "backend" / "stock_research.db"
WICS_CACHE_PATH = project_root / "data" / "wics_cache.json"
WICS_URL = "https://navercomp.wisereport.co.kr/v2/company/c1010001.aspx?cmp_cd={ticker}"
WICS_USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0 Safari/537.36"
)

TRACKED_FIELDS = (
    "name",
    "eval_date",
    "moat_score",
    "investment_value",
    "bm",
    "bigo_raw",
    "details",
)


def get_sector_key_from_bm(bm_text: str) -> str:
    """Return sector key for dashboard grouping.

    For broad bucket "제조업", use its first sub-sector for better readability.
    """
    if not bm_text:
        return ""

    parts = [p.strip() for p in str(bm_text).split("/") if p and p.strip()]
    if not parts:
        return ""

    top = parts[0]
    if top == "제조업" and len(parts) >= 2:
        return parts[1]
    return top


def normalize_ticker(raw_value) -> str:
    if raw_value is None:
        return ""
    text = str(raw_value).strip()
    if not text:
        return ""
    if text.endswith(".0"):
        text = text[:-2]
    digits = re.sub(r"[^0-9]", "", text)
    if digits and len(digits) <= 12:
        return digits.zfill(6) if len(digits) <= 6 else digits
    return text


def normalize_eval_date(raw_value) -> str:
    if raw_value is None:
        return ""
    if isinstance(raw_value, datetime):
        return raw_value.strftime("%Y-%m-%d")
    text = str(raw_value).strip()
    if not text:
        return ""
    return text.split(" ")[0]


def make_row_hash(stock: dict) -> str:
    payload = {
        "ticker": stock.get("ticker") or "",
        "name": stock.get("name") or "",
        "eval_date": stock.get("eval_date") or "",
        "moat_score": stock.get("moat_score"),
        "investment_value": stock.get("investment_value"),
        "bm": stock.get("bm") or "",
        "bigo_raw": stock.get("bigo_raw") or "",
        "details": stock.get("details") or {},
    }
    encoded = json.dumps(payload, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    return hashlib.sha256(encoded.encode("utf-8")).hexdigest()


def make_payload_hash(ticker: str, payload: dict) -> str:
    normalized = {
        "ticker": ticker or "",
        "name": payload.get("name") or "",
        "eval_date": payload.get("eval_date") or "",
        "moat_score": payload.get("moat_score"),
        "investment_value": payload.get("investment_value"),
        "bm": payload.get("bm") or "",
        "bigo_raw": payload.get("bigo_raw") or "",
        "details": payload.get("details") or {},
    }
    encoded = json.dumps(normalized, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    return hashlib.sha256(encoded.encode("utf-8")).hexdigest()


def extract_wics_from_html(html: str) -> Optional[str]:
    match = re.search(r"WICS\s*:\s*([^<\r\n]+)", html, flags=re.IGNORECASE)
    if not match:
        return None
    value = match.group(1).strip()
    return value or None


def load_wics_cache() -> Dict[str, Dict[str, str]]:
    if not WICS_CACHE_PATH.exists():
        return {}
    try:
        payload = json.loads(WICS_CACHE_PATH.read_text(encoding="utf-8"))
        return payload.get("items", {})
    except Exception:
        return {}


def save_wics_cache(items: Dict[str, Dict[str, str]]) -> None:
    payload = {"updated_at": datetime.now().isoformat(), "items": items}
    WICS_CACHE_PATH.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def fetch_wics_once(ticker: str, timeout: float = 8.0, retries: int = 2) -> Tuple[Optional[str], Optional[str]]:
    headers = {"User-Agent": WICS_USER_AGENT}
    url = WICS_URL.format(ticker=ticker)
    last_error = None
    for _ in range(max(1, retries)):
        try:
            resp = requests.get(url, headers=headers, timeout=timeout)
            if resp.status_code != 200:
                last_error = f"http_{resp.status_code}"
                continue
            wics = extract_wics_from_html(resp.text)
            if wics:
                return wics, None
            last_error = "wics_not_found"
        except Exception as exc:
            last_error = f"request_error:{exc.__class__.__name__}"
    return None, last_error


def resolve_wics_for_new_tickers(
    tickers: List[str],
    timeout: float = 8.0,
    retries: int = 2,
) -> Dict[str, Dict[str, str]]:
    cache = load_wics_cache()
    now = datetime.now().isoformat()
    resolved: Dict[str, Dict[str, str]] = {}

    for ticker in sorted({str(t).strip() for t in tickers if str(t).strip()}):
        cached = cache.get(ticker)
        if cached and cached.get("wics"):
            resolved[ticker] = {
                "wics": str(cached.get("wics")),
                "status": "cache",
                "error": "",
            }
            continue

        wics, err = fetch_wics_once(ticker=ticker, timeout=timeout, retries=retries)
        if wics:
            resolved[ticker] = {"wics": wics, "status": "fetched", "error": ""}
            cache[ticker] = {"wics": wics, "fetched_at": now}
        else:
            resolved[ticker] = {"wics": "", "status": "failed", "error": err or "unknown"}

    save_wics_cache(cache)
    return resolved


def parse_bigo(bigo_text):
    """Parse variable-length 비고 text into structured values.

    Works with 2-line, 3-line, or 10+ line formats.
    """
    if not bigo_text:
        return {}

    text = str(bigo_text).replace("\r\n", "\n").replace("\r", "\n").strip()
    lines = [line.strip() for line in text.split("\n") if line and line.strip()]
    if not lines:
        return {}

    result = {"raw_lines": lines}
    extras = []
    markers = []

    for line in lines:
        parts = [part.strip() for part in line.split("|") if part and part.strip()]
        if not parts:
            continue

        for part in parts:
            if part.lower().startswith("mgb."):
                markers.append(part)

            moat_line_match = re.search(r"해자\s*([0-5])\s*/\s*5", part)
            if moat_line_match:
                result["moat_line_score"] = int(moat_line_match.group(1))

            value_line_match = re.search(r"투자가치\s*([0-5])\s*/\s*5", part)
            if value_line_match:
                result["value_line_score"] = int(value_line_match.group(1))

            ev_match = re.search(r"증거\s*(\d+)\s*건\s*\(\s*q\s*([\d.]+)\s*\)", part)
            if ev_match:
                result["evidence_count"] = int(ev_match.group(1))
                result["evidence_quality"] = float(ev_match.group(2))
                continue

            growth_match = re.search(r"성장\s*([+-]\d+)", part)
            if growth_match:
                result["growth_adj"] = int(growth_match.group(1))
                if "성장" in part:
                    result["growth_reason"] = part
                continue

            rev_match = re.search(r"TTM매출\s*([-\d,.]+[조억만]?)\s*원?", part)
            if rev_match:
                result["ttm_revenue"] = rev_match.group(1)
                continue

            op_match = re.search(r"영업이익\s*([-\d,.]+[조억만]?)\s*원?\s*\(([-\d.]+%)\)", part)
            if op_match:
                result["op_income"] = op_match.group(1)
                result["op_margin"] = op_match.group(2)
                continue

            op_only_match = re.search(r"영업이익\s*([-\d,.]+[조억만]?)\s*원?", part)
            if op_only_match and "op_income" not in result:
                result["op_income"] = op_only_match.group(1)
                continue

            mult_match = re.search(r"op_multiple\s*([-\d.]+)\s*x", part, flags=re.IGNORECASE)
            if mult_match:
                result["op_multiple"] = float(mult_match.group(1))
                continue

            cap_match = re.search(r"시총\s*([-\d,.]+[조억만]?)\s*원?", part)
            if cap_match:
                result["market_cap"] = cap_match.group(1)
                continue

            price_match = re.search(r"현재가\s*([\d,]+)\s*원", part)
            if price_match:
                result["current_price"] = int(price_match.group(1).replace(",", ""))
                continue

            src_match = re.search(r"\[([^\]]+)\]", part)
            if src_match:
                result["data_source"] = src_match.group(1).strip()
                continue

            extras.append(part)

    if markers:
        result["markers"] = markers
    if extras:
        result["extra_parts"] = extras
    return result


def load_stocks_from_excel(xlsx_path: str = XLSX_PATH) -> Tuple[List[dict], int]:
    print(f"[EXTRACT] Loading: {xlsx_path}")
    print(f"[EXTRACT] This may take 1-2 minutes for large workbooks...")
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    # Build header map
    headers = {}
    for c in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=c).value
        if val:
            headers[str(val).strip()] = c

    col_code = headers.get("종목코드")
    col_name = headers.get("종목명")
    col_date = headers.get("평가일자")
    col_value = headers.get("투자가치")
    col_bm = headers.get("BM")
    col_moat = headers.get("해자")
    col_bigo = headers.get("비고")

    if not col_code or not col_name:
        print("[ERROR] Required columns not found:", headers)
        return [], 0

    stocks = []
    seen = {}
    duplicate_count = 0
    for row in range(2, ws.max_row + 1):
        ticker = normalize_ticker(ws.cell(row=row, column=col_code).value)
        name = str(ws.cell(row=row, column=col_name).value or "").strip()
        if not ticker or not name:
            continue

        moat_raw = ws.cell(row=row, column=col_moat).value if col_moat else None
        value_raw = ws.cell(row=row, column=col_value).value if col_value else None
        eval_date = normalize_eval_date(ws.cell(row=row, column=col_date).value) if col_date else ""
        bm = str(ws.cell(row=row, column=col_bm).value or "") if col_bm else ""
        bigo = ws.cell(row=row, column=col_bigo).value if col_bigo else None

        moat_int = None
        if moat_raw is not None:
            try:
                moat_int = int(float(str(moat_raw)))
            except (ValueError, TypeError):
                pass

        value_int = None
        if value_raw is not None:
            try:
                value_int = int(float(str(value_raw)))
            except (ValueError, TypeError):
                pass

        bigo_raw = str(bigo).strip() if bigo is not None else ""
        parsed = parse_bigo(bigo)

        stock_row = {
            "ticker": ticker,
            "name": name,
            "eval_date": eval_date,
            "moat_score": moat_int,
            "investment_value": value_int,
            "bm": bm,
            "bigo_raw": bigo_raw,
            "details": parsed,
        }
        if ticker in seen:
            duplicate_count += 1
        seen[ticker] = stock_row

    wb.close()
    stocks = list(seen.values())
    return stocks, duplicate_count


def build_dashboard_payload(stocks: List[dict], extracted_at: Optional[str] = None) -> dict:
    if extracted_at is None:
        extracted_at = datetime.now().isoformat()

    evaluated = [s for s in stocks if s["moat_score"] is not None]
    unevaluated = [s for s in stocks if s["moat_score"] is None]

    moat_dist = {str(i): 0 for i in range(6)}
    value_dist = {str(i): 0 for i in range(6)}
    matrix = {}
    for m in range(6):
        for v in range(6):
            matrix[f"{m},{v}"] = 0

    sectors = {}
    moat_sum = 0
    value_sum = 0

    for s in evaluated:
        m = s["moat_score"]
        v = s["investment_value"] if s["investment_value"] is not None else 0
        moat_dist[str(m)] = moat_dist.get(str(m), 0) + 1
        value_dist[str(v)] = value_dist.get(str(v), 0) + 1
        matrix[f"{m},{v}"] = matrix.get(f"{m},{v}", 0) + 1
        moat_sum += m
        value_sum += v

        sector = get_sector_key_from_bm(s["bm"])
        if sector and sector not in ("분석실패", "None", ""):
            sectors[sector] = sectors.get(sector, 0) + 1

    high_value = [s for s in evaluated if (s["investment_value"] or 0) >= 4]

    return {
        "meta": {
            "total_stocks": len(stocks),
            "evaluated_count": len(evaluated),
            "unevaluated_count": len(unevaluated),
            "avg_moat": round(moat_sum / len(evaluated), 2) if evaluated else 0,
            "avg_value": round(value_sum / len(evaluated), 2) if evaluated else 0,
            "high_value_count": len(high_value),
            "extracted_at": extracted_at,
        },
        "distributions": {
            "moat": moat_dist,
            "investment_value": value_dist,
            "matrix": matrix,
        },
        "sectors": sectors,
        "stocks": stocks,
    }


def _as_snapshot_payload(stock: dict) -> dict:
    return {
        "name": stock.get("name") or "",
        "eval_date": stock.get("eval_date") or "",
        "moat_score": stock.get("moat_score"),
        "investment_value": stock.get("investment_value"),
        "bm": stock.get("bm") or "",
        "bigo_raw": stock.get("bigo_raw") or "",
        "details": stock.get("details") or {},
    }


def _compute_changes(before_payload: dict, after_payload: dict) -> Dict[str, dict]:
    changes = {}
    for field in TRACKED_FIELDS:
        before_value = before_payload.get(field)
        after_value = after_payload.get(field)
        if before_value != after_value:
            changes[field] = {"before": before_value, "after": after_value}
    return changes


def sync_stocks_to_db(
    db: Session,
    stocks: List[dict],
    duplicate_count: int,
    source_file: str,
    imported_at: Optional[datetime] = None,
    verify_wics_on_insert: bool = True,
) -> dict:
    from backend.app.models.moat_data import MoatStockHistory, MoatStockSnapshot, MoatStockSyncRun

    imported_at = imported_at or datetime.utcnow()
    existing_rows = db.query(MoatStockSnapshot).all()
    existing_map = {row.ticker: row for row in existing_rows}
    new_tickers = [stock["ticker"] for stock in stocks if stock["ticker"] not in existing_map]
    wics_resolution = {}
    if verify_wics_on_insert and new_tickers:
        wics_resolution = resolve_wics_for_new_tickers(new_tickers)

    run = MoatStockSyncRun(
        source_file=source_file,
        started_at=imported_at,
        finished_at=imported_at,
        total_rows=len(stocks),
        duplicate_ticker_count=duplicate_count,
    )
    db.add(run)
    db.flush()

    inserted = 0
    updated = 0
    unchanged = 0
    reactivated = 0
    deactivated = 0
    errors = 0

    active_tickers = set()
    for stock in stocks:
        ticker = stock["ticker"]
        active_tickers.add(ticker)
        payload = _as_snapshot_payload(stock)

        existing = existing_map.get(ticker)
        if existing is None:
            wics_meta = {"status": "skipped", "wics": "", "error": "", "bm_before": payload["bm"]}
            if verify_wics_on_insert:
                resolved = wics_resolution.get(ticker, {})
                if resolved.get("wics"):
                    payload["bm"] = resolved["wics"]
                wics_meta = {
                    "status": resolved.get("status", "failed"),
                    "wics": resolved.get("wics", ""),
                    "error": resolved.get("error", ""),
                    "bm_before": wics_meta["bm_before"],
                    "bm_after": payload["bm"],
                }

            row_hash = make_payload_hash(ticker=ticker, payload=payload)
            snapshot = MoatStockSnapshot(
                ticker=ticker,
                name=payload["name"],
                eval_date=payload["eval_date"],
                moat_score=payload["moat_score"],
                investment_value=payload["investment_value"],
                bm=payload["bm"],
                bigo_raw=payload["bigo_raw"],
                details_json=json.dumps(payload["details"], ensure_ascii=False),
                source_file=source_file,
                source_row_hash=row_hash,
                source_updated_at=imported_at,
                is_active=True,
            )
            db.add(snapshot)
            db.add(
                MoatStockHistory(
                    run_id=run.id,
                    ticker=ticker,
                    action="insert",
                    change_json=json.dumps({"after": payload, "wics_verification": wics_meta}, ensure_ascii=False),
                    source_file=source_file,
                    source_row_hash=row_hash,
                    happened_at=imported_at,
                )
            )
            inserted += 1
            continue

        before_payload = {
            "name": existing.name or "",
            "eval_date": existing.eval_date or "",
            "moat_score": existing.moat_score,
            "investment_value": existing.investment_value,
            "bm": existing.bm or "",
            "bigo_raw": (existing.bigo_raw or "") if hasattr(existing, "bigo_raw") else "",
            "details": {},
        }
        try:
            before_payload["details"] = json.loads(existing.details_json or "{}")
        except json.JSONDecodeError:
            before_payload["details"] = {}
        was_active = bool(existing.is_active)
        changes = _compute_changes(before_payload, payload)
        content_changed = bool(changes)
        row_hash = make_payload_hash(ticker=ticker, payload=payload)

        if content_changed or not was_active:
            existing.name = payload["name"]
            existing.eval_date = payload["eval_date"]
            existing.moat_score = payload["moat_score"]
            existing.investment_value = payload["investment_value"]
            existing.bm = payload["bm"]
            existing.bigo_raw = payload["bigo_raw"]
            existing.details_json = json.dumps(payload["details"], ensure_ascii=False)
            existing.source_file = source_file
            existing.source_row_hash = row_hash
            existing.source_updated_at = imported_at
            existing.is_active = True

            if not was_active and not content_changed:
                action = "reactivate"
                reactivated += 1
            else:
                action = "update"
                updated += 1

            db.add(
                MoatStockHistory(
                    run_id=run.id,
                    ticker=ticker,
                    action=action,
                    change_json=json.dumps(changes if changes else {"reactivated": True}, ensure_ascii=False),
                    source_file=source_file,
                    source_row_hash=row_hash,
                    happened_at=imported_at,
                )
            )
        else:
            unchanged += 1

    for row in existing_rows:
        if row.ticker in active_tickers:
            continue
        if not row.is_active:
            continue
        row.is_active = False
        row.source_updated_at = imported_at
        db.add(
            MoatStockHistory(
                run_id=run.id,
                ticker=row.ticker,
                action="deactivate",
                change_json=json.dumps({"reason": "missing_in_excel"}, ensure_ascii=False),
                source_file=source_file,
                source_row_hash=row.source_row_hash or "",
                happened_at=imported_at,
            )
        )
        deactivated += 1

    run.finished_at = datetime.utcnow()
    run.inserted_count = inserted
    run.updated_count = updated
    run.unchanged_count = unchanged
    run.reactivated_count = reactivated
    run.deactivated_count = deactivated
    run.error_count = errors
    run.note = (
        "Excel is canonical source. DB keeps latest snapshot plus change history."
    )
    db.flush()

    return {
        "run_id": run.id,
        "total_rows": len(stocks),
        "inserted": inserted,
        "updated": updated,
        "unchanged": unchanged,
        "reactivated": reactivated,
        "deactivated": deactivated,
        "duplicate_tickers": duplicate_count,
        "errors": errors,
    }


def _load_db_session_factory():
    sys.path.insert(0, str(project_root))
    import os

    os.environ.setdefault("DATABASE_URL", f"sqlite:///{DEFAULT_DB_PATH.as_posix()}")
    os.environ.setdefault("DEBUG", "false")
    import backend.app.models  # noqa: F401
    from backend.app.database import Base, SessionLocal, engine

    Base.metadata.create_all(bind=engine)
    _ensure_moat_schema(engine)
    return SessionLocal


def _ensure_moat_schema(engine) -> None:
    inspector = sa_inspect(engine)
    if not inspector.has_table("moat_stock_snapshot"):
        return

    columns = {col["name"] for col in inspector.get_columns("moat_stock_snapshot")}
    statements = []
    if "bigo_raw" not in columns:
        statements.append("ALTER TABLE moat_stock_snapshot ADD COLUMN bigo_raw TEXT DEFAULT ''")

    if not statements:
        return

    with engine.begin() as conn:
        for stmt in statements:
            conn.execute(sa_text(stmt))


def save_json(output: dict, output_path: str = OUTPUT_PATH):
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    return output_path


def extract():
    stocks, duplicate_count = load_stocks_from_excel(XLSX_PATH)
    output = build_dashboard_payload(stocks)

    session_factory = _load_db_session_factory()
    with session_factory() as db:
        sync_stats = sync_stocks_to_db(
            db=db,
            stocks=stocks,
            duplicate_count=duplicate_count,
            source_file=str(Path(XLSX_PATH).name),
        )
        db.commit()

    save_json(output, OUTPUT_PATH)

    print(f"[EXTRACT] Total stocks: {len(stocks)}")
    print(f"[EXTRACT] Evaluated: {output['meta']['evaluated_count']}")
    print(f"[EXTRACT] Avg moat: {output['meta']['avg_moat']}")
    print(f"[EXTRACT] Avg value: {output['meta']['avg_value']}")
    print(f"[EXTRACT] High value (>=4): {output['meta']['high_value_count']}")
    print(f"[EXTRACT] Sectors: {len(output['sectors'])}")
    print(f"[EXTRACT] JSON Output: {OUTPUT_PATH}")
    print(
        "[EXTRACT] DB Sync: "
        f"run_id={sync_stats['run_id']} "
        f"inserted={sync_stats['inserted']} "
        f"updated={sync_stats['updated']} "
        f"unchanged={sync_stats['unchanged']} "
        f"reactivated={sync_stats['reactivated']} "
        f"deactivated={sync_stats['deactivated']} "
        f"duplicates={sync_stats['duplicate_tickers']}"
    )
    return {
        "meta": output["meta"],
        "sync": sync_stats,
        "output_path": OUTPUT_PATH,
    }


if __name__ == "__main__":
    extract()
